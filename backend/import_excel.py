import pandas as pd
import asyncio
from motor.motor_asyncio import AsyncIOMotorClient
from datetime import datetime
import os
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Color

load_dotenv()

# Connexió a MongoDB
MONGO_URL = os.getenv('MONGO_URL', 'mongodb://localhost:27017')
DB_NAME = os.getenv('DB_NAME', 'eltombdereus')
client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

def get_category_from_color(wb, row_index, col_index=0):
    """
    Detectar categoria segons el color de la cel·la
    Blau → Serveis
    Verd → Comerç
    Salmó/Rosa → Bellesa
    Taronja → Restauració
    """
    ws = wb.active
    cell = ws.cell(row=row_index + 2, column=col_index + 1)  # +2 perquè openpyxl comença a 1 i té header
    
    if cell.fill and cell.fill.start_color:
        color_hex = str(cell.fill.start_color.rgb) if hasattr(cell.fill.start_color, 'rgb') else str(cell.fill.start_color.index)
        
        # Detectar colors (aproximats)
        if color_hex:
            color_hex = color_hex.upper()
            
            # Blau → Serveis
            if any(blue in color_hex for blue in ['0000FF', '0070C0', '4472C4', '5B9BD5']):
                return 'Serveis'
            # Verd → Comerç
            elif any(green in color_hex for green in ['00FF00', '70AD47', '00B050', '92D050']):
                return 'Comerç'
            # Salmó/Rosa → Bellesa
            elif any(salmon in color_hex for salmon in ['FFC0CB', 'F4B084', 'E7E6E6', 'FABF8F']):
                return 'Bellesa'
            # Taronja → Restauració
            elif any(orange in color_hex for orange in ['FFA500', 'ED7D31', 'F4B084', 'C65911']):
                return 'Restauració'
    
    return None

async def import_establishments_from_excel(file_path, category_default=""):
    """
    Importar establiments des d'un fitxer Excel amb detecció de colors
    """
    try:
        # Llegir Excel amb openpyxl per obtenir colors
        wb = load_workbook(file_path)
        
        # Llegir dades amb pandas
        df = pd.read_excel(file_path)
        
        print(f"Llegint {file_path}...")
        print(f"Columnes: {list(df.columns)}")
        print(f"Total files: {len(df)}")
        
        imported = 0
        skipped = 0
        errors = 0
        
        # Buscar índex de la columna 'Nom'
        nom_col_index = 0
        for i, col in enumerate(df.columns):
            if 'nom' in str(col).lower():
                nom_col_index = i
                break
        
        for index, row in df.iterrows():
            try:
                # Extreure dades
                name = str(row.get('Nom', '') or row.get('nom', '') or row.get('Nom establiment', '')).strip()
                
                # Saltar files buides o invàlides
                if not name or name == 'nan' or name == '':
                    skipped += 1
                    continue
                
                # Verificar si ja existeix (per nom)
                existing = await db.establishments.find_one({"name": name})
                if existing:
                    print(f"  ⚠️  Ja existeix: {name}")
                    skipped += 1
                    continue
                
                # Detectar categoria pel color de la cel·la
                category = get_category_from_color(wb, index, nom_col_index)
                if not category:
                    category = category_default
                
                # Preparar document
                establishment = {
                    "name": name,
                    "address": str(row.get('Adreça', '') or row.get('adreça', '')).strip() if pd.notna(row.get('Adreça')) or pd.notna(row.get('adreça')) else None,
                    "category": category,
                    "phone": str(row.get('Telèfon', '') or row.get('telèfon', '') or row.get('Telèfon de contacte', '')).strip() if pd.notna(row.get('Telèfon')) or pd.notna(row.get('telèfon')) or pd.notna(row.get('Telèfon de contacte')) else None,
                    "email": str(row.get('E-mail', '') or row.get('e-mail', '') or row.get('Correu electrònic', '') or row.get('correu electrònic', '')).strip() if pd.notna(row.get('E-mail')) or pd.notna(row.get('e-mail')) or pd.notna(row.get('Correu electrònic')) or pd.notna(row.get('correu electrònic')) else None,
                    "website": str(row.get('Adreça web', '')).strip() if pd.notna(row.get('Adreça web')) else None,
                    "facebook": str(row.get('Facebook', '')).strip() if pd.notna(row.get('Facebook')) else None,
                    "description": str(row.get('Descripció', '')).strip() if pd.notna(row.get('Descripció')) else None,
                    "latitude": float(row.get('Latitud')) if pd.notna(row.get('Latitud')) and row.get('Latitud') != '' else None,
                    "longitude": float(row.get('Longitud')) if pd.notna(row.get('Longitud')) and row.get('Longitud') != '' else None,
                    "created_at": datetime.utcnow(),
                }
                
                # Netejar valors 'nan'
                establishment = {k: v for k, v in establishment.items() if v and str(v) != 'nan'}
                
                # Insertar
                result = await db.establishments.insert_one(establishment)
                print(f"  ✅ Importat: {name} → {category}")
                imported += 1
                
            except Exception as e:
                print(f"  ❌ Error a fila {index}: {str(e)}")
                errors += 1
                continue
        
        print(f"\n📊 RESUM:")
        print(f"  ✅ Importats: {imported}")
        print(f"  ⚠️  Saltats: {skipped}")
        print(f"  ❌ Errors: {errors}")
        
        return {"imported": imported, "skipped": skipped, "errors": errors}
        
    except Exception as e:
        print(f"❌ Error llegint l'Excel: {str(e)}")
        return {"error": str(e)}

async def main():
    """
    Funció principal per executar la importació
    """
    print("🚀 IMPORTACIÓ D'ESTABLIMENTS DES D'EXCEL\n")
    
    # Fitxers a importar
    files = [
        {
            "path": "/tmp/restauracio.xlsx",
            "category": "Restauració"
        },
        {
            "path": "/tmp/comerc_serveis.xlsx", 
            "category": "Comerç i Serveis"
        }
    ]
    
    total_imported = 0
    total_skipped = 0
    total_errors = 0
    
    for file_info in files:
        if os.path.exists(file_info['path']):
            print(f"\n📁 Processant: {file_info['path']}")
            print(f"   Categoria: {file_info['category']}\n")
            
            result = await import_establishments_from_excel(
                file_info['path'], 
                file_info['category']
            )
            
            if 'error' not in result:
                total_imported += result['imported']
                total_skipped += result['skipped']
                total_errors += result['errors']
        else:
            print(f"⚠️  Fitxer no trobat: {file_info['path']}")
    
    print(f"\n\n🎯 RESUM TOTAL:")
    print(f"  ✅ Total importats: {total_imported}")
    print(f"  ⚠️  Total saltats: {total_skipped}")
    print(f"  ❌ Total errors: {total_errors}")

if __name__ == "__main__":
    asyncio.run(main())
