#!/usr/bin/env python3
"""
Script per importar establiments des de l'Excel V04
Esborra tots els establiments existents i importa els nous
"""

import asyncio
import openpyxl
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv
from pathlib import Path
import os
from datetime import datetime

# Load environment variables
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
db_name = os.environ['DB_NAME']

def clean_coordinate(coord_value):
    """
    Converteix coordenades sense decimals a format correcte
    Ex: 4115663820 -> 41.15663820
        111013110 -> 1.11013110
    """
    if coord_value is None:
        return None
    
    try:
        # Convertir a string i eliminar espais
        coord_str = str(coord_value).strip().replace(' ', '')
        
        # Si ja té punt decimal, retornar com a float
        if '.' in coord_str:
            return float(coord_str)
        
        # Si és un número sense decimals
        if coord_str.isdigit():
            coord_int = int(coord_str)
            # Dividir per 10^8 per obtenir el decimal correcte
            return coord_int / 100000000
        
        # Intent conversió directa
        return float(coord_str)
    except (ValueError, TypeError):
        return None

def clean_string(value):
    """Neteja strings, retorna None si està buit"""
    if value is None:
        return None
    value = str(value).strip()
    return value if value else None

def clean_url(value):
    """Neteja URLs i afegeix https:// si cal"""
    if value is None:
        return None
    value = str(value).strip()
    if not value:
        return None
    
    # Si comença amb @ (Instagram), afegir URL completa
    if value.startswith('@'):
        return f"https://www.instagram.com/{value[1:]}/"
    
    # Si no té protocol, afegir https://
    if value and not value.startswith(('http://', 'https://')):
        return f"https://{value}"
    
    return value

async def import_establishments():
    """Importa establiments des de l'Excel"""
    
    print("🚀 Iniciant importació d'establiments...")
    
    # Connectar a MongoDB
    client = AsyncIOMotorClient(mongo_url)
    db = client[db_name]
    establishments_collection = db.establishments
    
    try:
        # 1. ESBORRAR TOTS els establiments existents
        print("\n⚠️  ESBORRANT tots els establiments existents...")
        result = await establishments_collection.delete_many({})
        print(f"   ✅ Esborrats {result.deleted_count} establiments")
        
        # 2. Carregar Excel
        print("\n📂 Carregant Excel...")
        wb = openpyxl.load_workbook('establiments_v04.xlsx')
        ws = wb.active
        
        # Obtenir headers
        headers = [cell.value for cell in ws[1]]
        print(f"   ✅ Headers detectats: {len(headers)} columnes")
        
        # 3. Processar i importar cada fila
        print("\n📥 Important establiments...")
        imported_count = 0
        error_count = 0
        
        for row_idx in range(2, ws.max_row + 1):
            try:
                # Extreure dades de la fila
                row_data = {}
                for col_idx, header in enumerate(headers, 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    row_data[header] = cell_value
                
                # Saltar files sense nom
                if not row_data.get('nom'):
                    continue
                
                # Crear document d'establiment
                establishment = {
                    'name': clean_string(row_data.get('nom')),
                    'status': clean_string(row_data.get('estatus')),
                    'address': clean_string(row_data.get('adreça')),
                    'latitude': clean_coordinate(row_data.get('latitut')),
                    'longitude': clean_coordinate(row_data.get('longitut')),
                    'phone': clean_string(row_data.get('telefon')),
                    'whatsapp': clean_string(row_data.get('whatsapp')),
                    'email': clean_string(row_data.get('e-mail')),
                    'vat_number': clean_string(row_data.get('vat_number')),
                    'nif': clean_string(row_data.get('vat_number')),  # Mateix camp
                    'iban': clean_string(row_data.get('IBAN')),
                    'collectiu': clean_string(row_data.get('collectiu')),
                    'quota': clean_string(row_data.get('quota')),
                    'description': clean_string(row_data.get('descripció')),
                    
                    # URLs
                    'logo_url': clean_url(row_data.get('logo_url')),
                    'image_url': clean_url(row_data.get('logo_url')),  # També al camp image_url
                    'imatge1_url': clean_url(row_data.get('imatge1_url')),
                    'imatge2_url': clean_url(row_data.get('imatge2_url')),
                    'web_url': clean_url(row_data.get('web_url')),
                    'website': clean_url(row_data.get('web_url')),  # També al camp website
                    'tourvirtual_url': clean_url(row_data.get('tourvirtual_url')),
                    
                    # Xarxes socials
                    'instagram': clean_url(row_data.get('instagram')),
                    'facebook': clean_url(row_data.get('facebook')),
                    'twitter': clean_url(row_data.get('twitter')),
                    
                    # Categories
                    'category': clean_string(row_data.get('categoria estadistica')),
                    'altres_categories': clean_string(row_data.get('altres categories')),
                    
                    # Participació
                    'lbac': clean_string(row_data.get('LBAC')),
                    'tiquets': clean_string(row_data.get('tiquets')),
                    'gimcana': clean_string(row_data.get('gimcana')),
                    'tombs': clean_string(row_data.get('TOMBS')),
                    
                    # Metadata
                    'establishment_type': 'local_associat',  # Per defecte
                    'visible_in_public_list': True,
                    'created_at': datetime.now(),
                    'updated_at': datetime.now()
                }
                
                # Eliminar camps None per no ocupar espai
                establishment = {k: v for k, v in establishment.items() if v is not None}
                
                # Inserir a MongoDB
                await establishments_collection.insert_one(establishment)
                imported_count += 1
                
                if imported_count % 50 == 0:
                    print(f"   ... {imported_count} establiments importats")
                
            except Exception as e:
                error_count += 1
                print(f"   ❌ Error a la fila {row_idx}: {e}")
        
        print(f"\n✅ Importació completada!")
        print(f"   📊 Total importats: {imported_count}")
        print(f"   ❌ Errors: {error_count}")
        
        # Verificar
        total_in_db = await establishments_collection.count_documents({})
        print(f"   🗄️  Total a la base de dades: {total_in_db}")
        
    except Exception as e:
        print(f"\n❌ Error durant la importació: {e}")
        raise
    finally:
        client.close()

if __name__ == "__main__":
    asyncio.run(import_establishments())
