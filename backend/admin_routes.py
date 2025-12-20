"""
Admin routes per gestionar continguts del backoffice
"""
from fastapi import APIRouter, HTTPException, Header, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from datetime import datetime
from typing import Optional, List
from pydantic import BaseModel, Field
from bson import ObjectId
import os
import pandas as pd
import tempfile
import base64
import io
from openpyxl import Workbook
from dotenv import load_dotenv
from pathlib import Path
from participation_tracker import (
    get_all_tags,
    get_users_by_tag,
    get_participation_stats
)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

admin_router = APIRouter(prefix="/admin", tags=["Admin"])

# Models de Rols
class RoleBase(BaseModel):
    name: str
    description: Optional[str] = None
    permissions: Optional[List[str]] = []
    color: Optional[str] = None  # Color per identificar visualment

class RoleCreate(RoleBase):
    pass

class RoleUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    permissions: Optional[List[str]] = None
    color: Optional[str] = None

class Role(RoleBase):
    id: str = Field(alias="_id")
    code: str  # Codi únic (ex: "admin", "user", "local_associat")
    is_system: bool = False  # No es pot eliminar si és del sistema
    created_at: datetime = Field(default_factory=datetime.utcnow)
    
    class Config:
        populate_by_name = True
        json_encoders = {ObjectId: str}

# Models
class NewsBase(BaseModel):
    title: str
    content: str
    author: Optional[str] = None
    image_url: Optional[str] = None
    category: Optional[str] = None
    published: bool = True

class NewsCreate(NewsBase):
    pass

class NewsUpdate(BaseModel):
    title: Optional[str] = None
    content: Optional[str] = None
    author: Optional[str] = None
    image_url: Optional[str] = None
    category: Optional[str] = None
    published: Optional[bool] = None

class News(NewsBase):
    id: str = Field(alias="_id")
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)
    
    class Config:
        populate_by_name = True
        json_encoders = {ObjectId: str}

class EstablishmentCreate(BaseModel):
    name: str
    description: Optional[str] = None
    category: Optional[str] = None
    subcategory: Optional[str] = None
    address: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    phone: Optional[str] = None
    whatsapp: Optional[str] = None
    website: Optional[str] = None
    image_url: Optional[str] = None
    email: Optional[str] = None
    nif: Optional[str] = None
    social_media: Optional[dict] = {}

class EstablishmentUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    subcategory: Optional[str] = None
    address: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    phone: Optional[str] = None
    whatsapp: Optional[str] = None
    website: Optional[str] = None
    image_url: Optional[str] = None
    email: Optional[str] = None
    nif: Optional[str] = None
    social_media: Optional[dict] = None

class OfferCreate(BaseModel):
    establishment_id: str
    title: str
    description: str
    discount: Optional[str] = None
    valid_from: datetime
    valid_until: datetime
    image_url: Optional[str] = None
    terms: Optional[str] = None
    web_link: Optional[str] = None
    phone: Optional[str] = None

class OfferUpdate(BaseModel):
    establishment_id: Optional[str] = None
    title: Optional[str] = None
    description: Optional[str] = None
    discount: Optional[str] = None
    valid_from: Optional[datetime] = None
    valid_until: Optional[datetime] = None
    image_url: Optional[str] = None
    terms: Optional[str] = None
    web_link: Optional[str] = None
    phone: Optional[str] = None

class EventCreate(BaseModel):
    establishment_id: Optional[str] = None
    title: str
    description: str
    discount: Optional[str] = None
    valid_from: datetime
    valid_until: datetime
    image_url: Optional[str] = None
    terms: Optional[str] = None
    web_link: Optional[str] = None
    phone: Optional[str] = None
    facebook_link: Optional[str] = None
    instagram_link: Optional[str] = None
    twitter_link: Optional[str] = None
    youtube_link: Optional[str] = None
    linkedin_link: Optional[str] = None
    tiktok_link: Optional[str] = None
    participating_establishments: Optional[List[str]] = []

class EventUpdate(BaseModel):
    establishment_id: Optional[str] = None
    title: Optional[str] = None
    description: Optional[str] = None
    discount: Optional[str] = None
    valid_from: Optional[datetime] = None
    valid_until: Optional[datetime] = None
    image_url: Optional[str] = None
    terms: Optional[str] = None
    web_link: Optional[str] = None
    phone: Optional[str] = None
    facebook_link: Optional[str] = None
    instagram_link: Optional[str] = None
    twitter_link: Optional[str] = None
    youtube_link: Optional[str] = None
    linkedin_link: Optional[str] = None
    tiktok_link: Optional[str] = None
    participating_establishments: Optional[List[str]] = None

class UserUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    birth_date: Optional[datetime] = None
    gender: Optional[str] = None
    location: Optional[str] = None
    role: Optional[str] = None  # Mantenim per compatibilitat
    roles: Optional[List[str]] = None  # Nou camp per múltiples rols

class ImageUploadResponse(BaseModel):
    url: str
    filename: str

# Middleware per verificar si l'usuari és admin
async def verify_admin(authorization: str = Header(None)):
    if not authorization:
        raise HTTPException(status_code=401, detail="No autoritzat")
    
    # Suportar diferents formats de token
    token = authorization.replace("Bearer ", "").replace("token_", "")
    
    try:
        # Primer intentar buscar per token directament
        user = await db.users.find_one({"token": token})
        if user:
            if user.get('role') != 'admin':
                raise HTTPException(status_code=403, detail="Accés denegat - només administradors")
            return user
        
        # Si no es troba, intentar buscar per _id
        try:
            user = await db.users.find_one({"_id": ObjectId(token)})
            if user:
                if user.get('role') != 'admin':
                    raise HTTPException(status_code=403, detail="Accés denegat - només administradors")
                return user
        except:
            pass
    except Exception:
        pass
    
    raise HTTPException(status_code=401, detail="Token invàlid")

# ============================================================================
# ESTABLIMENTS - Admin CRUD
# ============================================================================

@admin_router.post("/establishments")
async def create_establishment(
    establishment: EstablishmentCreate,
    authorization: str = Header(None)
):
    """Crear un nou establiment"""
    await verify_admin(authorization)
    
    establishment_dict = establishment.dict()
    establishment_dict['created_at'] = datetime.utcnow()
    establishment_dict['updated_at'] = datetime.utcnow()
    
    result = await db.establishments.insert_one(establishment_dict)
    establishment_dict['_id'] = str(result.inserted_id)
    
    return establishment_dict

@admin_router.post("/establishments/{establishment_id}/assign-owner")
async def assign_owner_to_establishment(
    establishment_id: str,
    owner_email: str,
    authorization: str = Header(None)
):
    """Assignar un propietari a un establiment per email"""
    await verify_admin(authorization)
    
    # Buscar l'usuari per email
    user = await db.users.find_one({"email": owner_email})
    
    if not user:
        raise HTTPException(status_code=404, detail=f"Usuari amb email {owner_email} no trobat")
    
    # Actualitzar l'establiment amb l'owner_id
    result = await db.establishments.update_one(
        {"_id": ObjectId(establishment_id)},
        {"$set": {
            "owner_id": user['_id'],
            "updated_at": datetime.utcnow()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Establiment no trobat")
    
    # Actualitzar l'usuari al rol local_associat si no ho és
    if user.get('role') != 'local_associat' and user.get('role') != 'admin':
        await db.users.update_one(
            {"_id": user['_id']},
            {"$set": {"role": "local_associat"}}
        )
    
    updated = await db.establishments.find_one({"_id": ObjectId(establishment_id)})
    updated['_id'] = str(updated['_id'])
    # Convertir owner_id a string si existeix
    if updated.get('owner_id'):
        updated['owner_id'] = str(updated['owner_id'])
    updated['owner_email'] = owner_email
    
    return {
        "success": True,
        "message": f"Establiment assignat a {owner_email}",
        "establishment": updated
    }

@admin_router.put("/establishments/{establishment_id}")
async def update_establishment(
    establishment_id: str,
    establishment: EstablishmentUpdate,
    authorization: str = Header(None)
):
    """Actualitzar un establiment"""
    await verify_admin(authorization)
    
    update_data = {k: v for k, v in establishment.dict().items() if v is not None}
    update_data['updated_at'] = datetime.utcnow()
    
    result = await db.establishments.update_one(
        {"_id": ObjectId(establishment_id)},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Establiment no trobat")
    
    updated = await db.establishments.find_one({"_id": ObjectId(establishment_id)})
    updated['_id'] = str(updated['_id'])
    
    # Obtenir email del propietari si en té
    if updated.get('owner_id'):
        owner = await db.users.find_one({"_id": updated['owner_id']})
        if owner:
            updated['owner_email'] = owner.get('email')
    
    return updated

@admin_router.get("/establishments")
async def get_all_establishments_admin(
    authorization: str = Header(None)
):
    """Obtenir TOTS els establiments (incloent tancats) per a administradors"""
    await verify_admin(authorization)
    
    # Retornar tots els establiments sense filtrar per visible_in_public_list
    establishments = await db.establishments.find({}).to_list(1000)
    
    for est in establishments:
        est['_id'] = str(est['_id'])
        est['id'] = est['_id']
        if est.get('owner_id'):
            est['owner_id'] = str(est['owner_id'])
    
    return establishments

@admin_router.get("/establishments/{establishment_id}/owner")
async def get_establishment_owner(
    establishment_id: str,
    authorization: str = Header(None)
):
    """Obtenir el propietari d'un establiment"""
    await verify_admin(authorization)
    
    establishment = await db.establishments.find_one({"_id": ObjectId(establishment_id)})
    
    if not establishment:
        raise HTTPException(status_code=404, detail="Establiment no trobat")
    
    if not establishment.get('owner_id'):
        return {"owner": None, "message": "Aquest establiment no té propietari assignat"}
    
    owner = await db.users.find_one({"_id": establishment['owner_id']})
    
    if not owner:
        return {"owner": None, "message": "Propietari no trobat"}
    
    return {
        "owner": {
            "id": str(owner['_id']),
            "name": owner.get('name'),
            "email": owner.get('email'),
            "role": owner.get('role')
        }
    }

@admin_router.delete("/establishments/{establishment_id}")
async def delete_establishment(
    establishment_id: str,
    authorization: str = Header(None)
):
    """Eliminar un establiment"""
    await verify_admin(authorization)
    
    result = await db.establishments.delete_one({"_id": ObjectId(establishment_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Establiment no trobat")
    
    return {"success": True, "message": "Establiment eliminat"}

@admin_router.put("/establishments/{establishment_id}/assign-owner")
async def assign_establishment_owner(
    establishment_id: str,
    user_id: Optional[str] = None,
    authorization: str = Header(None)
):
    """Assignar o desassignar un propietari a un establiment"""
    await verify_admin(authorization)
    
    # Verificar que l'establiment existeix
    establishment = await db.establishments.find_one({"_id": ObjectId(establishment_id)})
    if not establishment:
        raise HTTPException(status_code=404, detail="Establiment no trobat")
    
    # Si user_id és None, desassignem el propietari
    if user_id is None:
        await db.establishments.update_one(
            {"_id": ObjectId(establishment_id)},
            {"$set": {"owner_id": None, "updated_at": datetime.utcnow()}}
        )
        return {"success": True, "message": "Propietari desassignat correctament"}
    
    # Verificar que l'usuari existeix
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="Usuari no trobat")
    
    # Permetre assignar qualsevol usuari registrat
    # (l'admin pot decidir qui és propietari d'un establiment)
    
    # Assignar el propietari
    await db.establishments.update_one(
        {"_id": ObjectId(establishment_id)},
        {"$set": {"owner_id": ObjectId(user_id), "updated_at": datetime.utcnow()}}
    )
    
    return {
        "success": True, 
        "message": "Propietari assignat correctament",
        "owner": {
            "id": str(user['_id']),
            "name": user.get('name'),
            "email": user.get('email')
        }
    }

@admin_router.get("/users/local-associats")
async def get_local_associats(
    authorization: str = Header(None),
    email: str = None
):
    """Obtenir tots els usuaris (filtrable per email per cerca)"""
    await verify_admin(authorization)
    
    # Si es proporciona un email, cercar només aquest usuari
    if email:
        user = await db.users.find_one({"email": email.lower()})
        if user:
            return [{
                "id": str(user['_id']),
                "name": user.get('name', ''),
                "email": user.get('email', ''),
                "role": user.get('role', 'user')
            }]
        return []
    
    # Retornar TOTS els usuaris registrats (no només local_associat)
    users = await db.users.find({}).to_list(1000)
    
    result = []
    for user in users:
        result.append({
            "id": str(user['_id']),
            "name": user.get('name', ''),
            "email": user.get('email', ''),
            "role": user.get('role', 'user')
        })
    
    return result


# ============================================================================
# OFERTES - Admin CRUD
# ============================================================================

@admin_router.post("/offers")
async def create_offer(
    offer: OfferCreate,
    authorization: str = Header(None)
):
    """Crear una nova oferta"""
    await verify_admin(authorization)
    
    offer_dict = offer.dict()
    offer_dict['created_at'] = datetime.utcnow()
    offer_dict['updated_at'] = datetime.utcnow()
    
    # Assegurar que les dates tenen hora si només són dates
    if 'valid_from' in offer_dict and offer_dict['valid_from']:
        if isinstance(offer_dict['valid_from'], str):
            if len(offer_dict['valid_from']) == 10:
                offer_dict['valid_from'] = offer_dict['valid_from'] + 'T00:00:00'
            try:
                offer_dict['valid_from'] = datetime.fromisoformat(offer_dict['valid_from'].replace('Z', '+00:00'))
            except:
                from dateutil import parser
                offer_dict['valid_from'] = parser.parse(offer_dict['valid_from'])
            
    if 'valid_until' in offer_dict and offer_dict['valid_until']:
        if isinstance(offer_dict['valid_until'], str):
            if len(offer_dict['valid_until']) == 10:
                offer_dict['valid_until'] = offer_dict['valid_until'] + 'T23:59:59'
            try:
                offer_dict['valid_until'] = datetime.fromisoformat(offer_dict['valid_until'].replace('Z', '+00:00'))
            except:
                from dateutil import parser
                offer_dict['valid_until'] = parser.parse(offer_dict['valid_until'])
    
    result = await db.offers.insert_one(offer_dict)
    offer_dict['_id'] = str(result.inserted_id)
    
    return offer_dict

@admin_router.put("/offers/{offer_id}")
async def update_offer(
    offer_id: str,
    offer: OfferUpdate,
    authorization: str = Header(None)
):
    """Actualitzar una oferta"""
    await verify_admin(authorization)
    
    update_data = {k: v for k, v in offer.dict().items() if v is not None}
    update_data['updated_at'] = datetime.utcnow()
    
    # Assegurar que les dates tenen hora si només són dates
    if 'valid_from' in update_data and update_data['valid_from']:
        if isinstance(update_data['valid_from'], str):
            if len(update_data['valid_from']) == 10:
                update_data['valid_from'] = update_data['valid_from'] + 'T00:00:00'
            try:
                update_data['valid_from'] = datetime.fromisoformat(update_data['valid_from'].replace('Z', '+00:00'))
            except:
                from dateutil import parser
                update_data['valid_from'] = parser.parse(update_data['valid_from'])
            
    if 'valid_until' in update_data and update_data['valid_until']:
        if isinstance(update_data['valid_until'], str):
            if len(update_data['valid_until']) == 10:
                update_data['valid_until'] = update_data['valid_until'] + 'T23:59:59'
            try:
                update_data['valid_until'] = datetime.fromisoformat(update_data['valid_until'].replace('Z', '+00:00'))
            except:
                from dateutil import parser
                update_data['valid_until'] = parser.parse(update_data['valid_until'])
    
    result = await db.offers.update_one(
        {"_id": ObjectId(offer_id)},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Oferta no trobada")
    
    updated = await db.offers.find_one({"_id": ObjectId(offer_id)})
    updated['_id'] = str(updated['_id'])
    
    return updated

@admin_router.delete("/offers/{offer_id}")
async def delete_offer(
    offer_id: str,
    authorization: str = Header(None)
):
    """Eliminar una oferta"""
    await verify_admin(authorization)
    
    result = await db.offers.delete_one({"_id": ObjectId(offer_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Oferta no trobada")
    
    return {"success": True, "message": "Oferta eliminada"}

# ============================================================================
# ESDEVENIMENTS - Admin CRUD
# ============================================================================

@admin_router.get("/events")
async def get_all_events(
    authorization: str = Header(None)
):
    """Obtenir tots els esdeveniments (admin)"""
    await verify_admin(authorization)
    
    # Només retornar esdeveniments amb l'estructura nova (valid_from i valid_until)
    events_cursor = db.events.find({
        "valid_from": {"$exists": True},
        "valid_until": {"$exists": True}
    }).sort("valid_from", -1)
    events = await events_cursor.to_list(length=None)
    
    for event in events:
        event['_id'] = str(event['_id'])
        event['id'] = str(event['_id'])
        if event.get('establishment_id'):
            event['establishment_id'] = str(event['establishment_id'])
    
    return events

@admin_router.post("/events")
async def create_event(
    event: EventCreate,
    authorization: str = Header(None)
):
    """Crear un nou esdeveniment"""
    await verify_admin(authorization)
    
    event_dict = event.dict()
    event_dict['created_at'] = datetime.utcnow()
    event_dict['updated_at'] = datetime.utcnow()
    
    # Assegurar que les dates tenen hora si només són dates
    if 'valid_from' in event_dict and event_dict['valid_from']:
        if isinstance(event_dict['valid_from'], str):
            # Si és format YYYY-MM-DD (del input HTML5), afegir hora 00:00:00
            if len(event_dict['valid_from']) == 10:  # Format YYYY-MM-DD
                event_dict['valid_from'] = event_dict['valid_from'] + 'T00:00:00'
            # Convertir a datetime
            try:
                event_dict['valid_from'] = datetime.fromisoformat(event_dict['valid_from'].replace('Z', '+00:00'))
            except:
                from dateutil import parser
                event_dict['valid_from'] = parser.parse(event_dict['valid_from'])
            
    if 'valid_until' in event_dict and event_dict['valid_until']:
        if isinstance(event_dict['valid_until'], str):
            # Si és format YYYY-MM-DD (del input HTML5), afegir hora 23:59:59
            if len(event_dict['valid_until']) == 10:  # Format YYYY-MM-DD
                event_dict['valid_until'] = event_dict['valid_until'] + 'T23:59:59'
            # Convertir a datetime
            try:
                event_dict['valid_until'] = datetime.fromisoformat(event_dict['valid_until'].replace('Z', '+00:00'))
            except:
                from dateutil import parser
                event_dict['valid_until'] = parser.parse(event_dict['valid_until'])
    
    result = await db.events.insert_one(event_dict)
    event_id = str(result.inserted_id)
    event_dict['_id'] = event_id
    
    # Crear marcador automàticament amb el nom de l'esdeveniment
    tag_name = event_dict.get('title', 'Esdeveniment')
    await db.tags.update_one(
        {"name": tag_name},
        {
            "$setOnInsert": {
                "name": tag_name,
                "source_type": "event",
                "source_id": event_id,
                "description": f"Marcador creat automàticament per l'esdeveniment: {tag_name}",
                "created_at": datetime.utcnow(),
                "user_count": 0
            }
        },
        upsert=True
    )
    
    return event_dict

@admin_router.put("/events/{event_id}")
async def update_event(
    event_id: str,
    event: EventUpdate,
    authorization: str = Header(None)
):
    """Actualitzar un esdeveniment"""
    await verify_admin(authorization)
    
    update_data = {k: v for k, v in event.dict().items() if v is not None}
    update_data['updated_at'] = datetime.utcnow()
    
    # Assegurar que les dates tenen hora si només són dates
    if 'valid_from' in update_data and update_data['valid_from']:
        if isinstance(update_data['valid_from'], str):
            # Si és format YYYY-MM-DD (del input HTML5), afegir hora 00:00:00
            if len(update_data['valid_from']) == 10:  # Format YYYY-MM-DD
                update_data['valid_from'] = update_data['valid_from'] + 'T00:00:00'
            # Convertir a datetime
            try:
                update_data['valid_from'] = datetime.fromisoformat(update_data['valid_from'].replace('Z', '+00:00'))
            except:
                from dateutil import parser
                update_data['valid_from'] = parser.parse(update_data['valid_from'])
            
    if 'valid_until' in update_data and update_data['valid_until']:
        if isinstance(update_data['valid_until'], str):
            # Si és format YYYY-MM-DD (del input HTML5), afegir hora 23:59:59
            if len(update_data['valid_until']) == 10:  # Format YYYY-MM-DD
                update_data['valid_until'] = update_data['valid_until'] + 'T23:59:59'
            # Convertir a datetime
            try:
                update_data['valid_until'] = datetime.fromisoformat(update_data['valid_until'].replace('Z', '+00:00'))
            except:
                from dateutil import parser
                update_data['valid_until'] = parser.parse(update_data['valid_until'])
    
    result = await db.events.update_one(
        {"_id": ObjectId(event_id)},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Esdeveniment no trobat")
    
    updated = await db.events.find_one({"_id": ObjectId(event_id)})
    updated['_id'] = str(updated['_id'])
    
    return updated

@admin_router.delete("/events/{event_id}")
async def delete_event(
    event_id: str,
    authorization: str = Header(None)
):
    """Eliminar un esdeveniment"""
    await verify_admin(authorization)
    
    result = await db.events.delete_one({"_id": ObjectId(event_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Esdeveniment no trobat")
    
    return {"success": True, "message": "Esdeveniment eliminat"}

# ============================================================================
# NOTÍCIES - Admin CRUD
# ============================================================================

@admin_router.get("/news")
async def get_all_news(
    authorization: str = Header(None),
    skip: int = 0,
    limit: int = 50
):
    """Obtenir totes les notícies (incloses no publicades)"""
    await verify_admin(authorization)
    
    news = await db.news.find().sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    for item in news:
        item['_id'] = str(item['_id'])
    
    return news

@admin_router.post("/news")
async def create_news(
    news: NewsCreate,
    authorization: str = Header(None)
):
    """Crear una nova notícia"""
    await verify_admin(authorization)
    
    news_dict = news.dict()
    news_dict['created_at'] = datetime.utcnow()
    news_dict['updated_at'] = datetime.utcnow()
    
    result = await db.news.insert_one(news_dict)
    news_dict['_id'] = str(result.inserted_id)
    
    return news_dict

@admin_router.put("/news/{news_id}")
async def update_news(
    news_id: str,
    news: NewsUpdate,
    authorization: str = Header(None)
):
    """Actualitzar una notícia"""
    await verify_admin(authorization)
    
    update_data = {k: v for k, v in news.dict().items() if v is not None}
    update_data['updated_at'] = datetime.utcnow()
    
    result = await db.news.update_one(
        {"_id": ObjectId(news_id)},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Notícia no trobada")
    
    updated = await db.news.find_one({"_id": ObjectId(news_id)})
    updated['_id'] = str(updated['_id'])
    
    return updated

@admin_router.delete("/news/{news_id}")
async def delete_news(
    news_id: str,
    authorization: str = Header(None)
):
    """Eliminar una notícia"""
    await verify_admin(authorization)
    
    result = await db.news.delete_one({"_id": ObjectId(news_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Notícia no trobada")
    
    return {"success": True, "message": "Notícia eliminada"}

# ============================================================================
# USUARIS - Admin gestió
# ============================================================================

@admin_router.get("/users")
async def get_all_users(
    authorization: str = Header(None),
    skip: int = 0,
    limit: int = 100
):
    """Obtenir tots els usuaris"""
    await verify_admin(authorization)
    
    users = await db.users.find().skip(skip).limit(limit).to_list(limit)
    for user in users:
        user['id'] = str(user['_id'])
        user['_id'] = str(user['_id'])
        # No retornar password
        user.pop('password', None)
    
    return users

@admin_router.post("/users/create")
async def create_user_with_establishment(
    email: str,
    name: str,
    role: str = "local_associat",
    establishment_id: Optional[str] = None,
    authorization: str = Header(None)
):
    """Crear un nou usuari amb contrasenya automàtica i assignar establiment"""
    await verify_admin(authorization)
    
    from passlib.context import CryptContext
    import secrets
    import string
    
    pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
    
    # Verificar que l'email no existeix
    existing_user = await db.users.find_one({"email": email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Aquest email ja està registrat")
    
    # Generar contrasenya automàtica (8 caràcters: lletres i números)
    alphabet = string.ascii_letters + string.digits
    auto_password = ''.join(secrets.choice(alphabet) for i in range(8))
    
    # Hashear la contrasenya
    hashed_password = pwd_context.hash(auto_password)
    
    # Crear l'usuari
    new_user = {
        "email": email,
        "name": name,
        "role": role,
        "hashed_password": hashed_password,
        "phone": "",
        "birth_date": None,
        "gender": "",
        "address": "",
        "city": "",
        "language": "ca",
        "tags": [],
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow()
    }
    
    result = await db.users.insert_one(new_user)
    user_id = result.inserted_id
    
    # Si s'ha especificat un establiment, assignar-lo
    if establishment_id and role == "local_associat":
        await db.establishments.update_one(
            {"_id": ObjectId(establishment_id)},
            {"$set": {
                "owner_id": user_id,
                "updated_at": datetime.utcnow()
            }}
        )
    
    return {
        "success": True,
        "user_id": str(user_id),
        "email": email,
        "name": name,
        "role": role,
        "password": auto_password,  # Retornar la contrasenya per mostrar-la a l'admin
        "establishment_assigned": establishment_id is not None
    }

@admin_router.get("/users/{user_id}")
async def get_user_details(
    user_id: str,
    authorization: str = Header(None)
):
    """Obtenir detalls d'un usuari"""
    await verify_admin(authorization)
    
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="Usuari no trobat")
    
    user['_id'] = str(user['_id'])
    user.pop('password', None)
    
    return user

@admin_router.put("/users/{user_id}")
async def update_user(
    user_id: str,
    user_update: UserUpdate,
    authorization: str = Header(None)
):
    """Actualitzar un usuari"""
    await verify_admin(authorization)
    
    update_data = {k: v for k, v in user_update.dict().items() if v is not None}
    update_data['updated_at'] = datetime.utcnow()
    
    # Si s'envia roles (múltiples), actualitzem també role (singular) per compatibilitat
    if 'roles' in update_data and update_data['roles']:
        # Guardem el primer rol com a rol principal per compatibilitat
        update_data['role'] = update_data['roles'][0]
    # Si s'envia role (singular), convertim a array per roles
    elif 'role' in update_data and update_data['role']:
        if 'roles' not in update_data:
            update_data['roles'] = [update_data['role']]
    
    result = await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Usuari no trobat")
    
    updated = await db.users.find_one({"_id": ObjectId(user_id)})
    updated['_id'] = str(updated['_id'])
    updated.pop('password', None)
    
    return updated

@admin_router.delete("/users/{user_id}")
async def delete_user(
    user_id: str,
    authorization: str = Header(None)
):
    """Eliminar un usuari"""
    admin_user = await verify_admin(authorization)
    
    # No es pot auto-eliminar
    if str(admin_user['_id']) == user_id:
        raise HTTPException(status_code=400, detail="No pots eliminar el teu propi compte")
    
    result = await db.users.delete_one({"_id": ObjectId(user_id)})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Usuari no trobat")
    
    return {"success": True, "message": "Usuari eliminat correctament"}

# ============================================================================
# ROLS - Admin gestió
# ============================================================================

@admin_router.get("/roles")
async def get_all_roles(authorization: str = Header(None)):
    """Obtenir tots els rols"""
    await verify_admin(authorization)
    
    roles = await db.roles.find().to_list(100)
    for role in roles:
        role['_id'] = str(role['_id'])
    
    return roles

@admin_router.post("/roles")
async def create_role(
    role: RoleCreate,
    authorization: str = Header(None)
):
    """Crear un nou rol"""
    await verify_admin(authorization)
    
    # Generar codi únic a partir del nom
    code = role.name.lower().replace(' ', '_').replace('à', 'a').replace('è', 'e').replace('é', 'e').replace('í', 'i').replace('ò', 'o').replace('ó', 'o').replace('ú', 'u')
    
    # Verificar que no existeixi
    existing = await db.roles.find_one({"code": code})
    if existing:
        raise HTTPException(status_code=400, detail="Ja existeix un rol amb aquest nom")
    
    role_dict = role.dict()
    role_dict['code'] = code
    role_dict['is_system'] = False
    role_dict['created_at'] = datetime.utcnow()
    
    result = await db.roles.insert_one(role_dict)
    role_dict['_id'] = str(result.inserted_id)
    
    return role_dict

@admin_router.put("/roles/{role_id}")
async def update_role(
    role_id: str,
    role_update: RoleUpdate,
    authorization: str = Header(None)
):
    """Actualitzar un rol"""
    await verify_admin(authorization)
    
    # Verificar que no sigui un rol del sistema
    role = await db.roles.find_one({"_id": ObjectId(role_id)})
    if not role:
        raise HTTPException(status_code=404, detail="Rol no trobat")
    
    if role.get('is_system', False):
        raise HTTPException(status_code=400, detail="No es poden modificar els rols del sistema")
    
    update_data = {k: v for k, v in role_update.dict().items() if v is not None}
    
    # Si canvia el nom, actualitzar el codi
    if 'name' in update_data:
        code = update_data['name'].lower().replace(' ', '_').replace('à', 'a').replace('è', 'e').replace('é', 'e').replace('í', 'i').replace('ò', 'o').replace('ó', 'o').replace('ú', 'u')
        update_data['code'] = code
    
    result = await db.roles.update_one(
        {"_id": ObjectId(role_id)},
        {"$set": update_data}
    )
    
    updated = await db.roles.find_one({"_id": ObjectId(role_id)})
    updated['_id'] = str(updated['_id'])
    
    return updated

@admin_router.delete("/roles/{role_id}")
async def delete_role(
    role_id: str,
    authorization: str = Header(None)
):
    """Eliminar un rol"""
    await verify_admin(authorization)
    
    # Verificar que no sigui un rol del sistema
    role = await db.roles.find_one({"_id": ObjectId(role_id)})
    if not role:
        raise HTTPException(status_code=404, detail="Rol no trobat")
    
    if role.get('is_system', False):
        raise HTTPException(status_code=400, detail="No es poden eliminar els rols del sistema")
    
    # Verificar que no hi hagi usuaris amb aquest rol
    users_with_role = await db.users.count_documents({"role": role['code']})
    if users_with_role > 0:
        raise HTTPException(
            status_code=400,
            detail=f"No es pot eliminar el rol perquè hi ha {users_with_role} usuari(s) assignat(s)"
        )
    
    await db.roles.delete_one({"_id": ObjectId(role_id)})
    
    return {"success": True, "message": "Rol eliminat correctament"}

# ============================================================================
# TARGETES REGAL - Admin gestió
# ============================================================================

@admin_router.get("/gift-cards")
async def get_all_gift_cards(
    authorization: str = Header(None),
    skip: int = 0,
    limit: int = 100
):
    """Obtenir totes les targetes regal"""
    await verify_admin(authorization)
    
    gift_cards = await db.gift_cards.find().sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    for gc in gift_cards:
        gc['_id'] = str(gc['_id'])
    
    return gift_cards

# ============================================================================
# PUJADA D'IMATGES
# ============================================================================

@admin_router.post("/upload-image", response_model=ImageUploadResponse)
async def upload_image(
    file: UploadFile = File(...),
    authorization: str = Header(None)
):
    """Pujar una imatge (com a base64 per simplicitat)"""
    await verify_admin(authorization)
    
    # Llegir fitxer
    contents = await file.read()
    
    # Convertir a base64
    base64_image = base64.b64encode(contents).decode('utf-8')
    
    # Determinar el mime type
    content_type = file.content_type or 'image/jpeg'
    
    # Crear data URL
    data_url = f"data:{content_type};base64,{base64_image}"
    
    # Guardar a MongoDB per referència futura
    image_doc = {
        "filename": file.filename,
        "content_type": content_type,
        "data_url": data_url,
        "uploaded_at": datetime.utcnow()
    }
    
    result = await db.images.insert_one(image_doc)
    
    return ImageUploadResponse(
        url=data_url,
        filename=file.filename
    )


@admin_router.post("/local-associat/upload-image", response_model=ImageUploadResponse)
async def upload_image_local(
    file: UploadFile = File(...),
    authorization: str = Header(None)
):
    """Pujar una imatge (accessible per locals associats i admins)"""
    if not authorization:
        raise HTTPException(status_code=401, detail="Token no proporcionat")
    
    token = authorization.replace('Bearer ', '')
    user = await db.users.find_one({"token": token})
    
    if not user:
        raise HTTPException(status_code=401, detail="Usuari no trobat")
    
    if user.get('role') not in ['local_associat', 'admin']:
        raise HTTPException(status_code=403, detail="Accés denegat")
    
    # Llegir fitxer
    contents = await file.read()
    
    # Convertir a base64
    base64_image = base64.b64encode(contents).decode('utf-8')
    
    # Determinar el mime type
    content_type = file.content_type or 'image/jpeg'
    
    # Crear data URL
    data_url = f"data:{content_type};base64,{base64_image}"
    
    # Guardar a MongoDB per referència futura
    image_doc = {
        "filename": file.filename,
        "content_type": content_type,
        "data_url": data_url,
        "uploaded_at": datetime.utcnow(),
        "uploaded_by": str(user['_id'])
    }
    
    result = await db.images.insert_one(image_doc)
    
    return ImageUploadResponse(
        url=data_url,
        filename=file.filename
    )


# ============================================================================
# ESTADÍSTIQUES
# ============================================================================

def get_category_from_cell_color(wb, row_index, col_index=0):
    """Detectar categoria segons color de cel·la"""
    from openpyxl import load_workbook
    
    ws = wb.active
    cell = ws.cell(row=row_index + 2, column=col_index + 1)
    
    if cell.fill and cell.fill.start_color:
        color_hex = str(cell.fill.start_color.rgb) if hasattr(cell.fill.start_color, 'rgb') else str(cell.fill.start_color.index)
        
        if color_hex:
            color_hex = color_hex.upper()
            
            # Detectar colors
            if any(blue in color_hex for blue in ['0000FF', '0070C0', '4472C4', '5B9BD5']):
                return 'Serveis'
            elif any(green in color_hex for green in ['00FF00', '70AD47', '00B050', '92D050']):
                return 'Comerç'
            elif any(salmon in color_hex for salmon in ['FFC0CB', 'F4B084', 'E7E6E6', 'FABF8F']):
                return 'Bellesa'
            elif any(orange in color_hex for orange in ['FFA500', 'ED7D31', 'F4B084', 'C65911']):
                return 'Restauració'
    
    return None

@admin_router.post("/import-establishments")
async def import_establishments_excel(
    file: UploadFile = File(...),
    authorization: str = Header(None),
    category: str = Form("")
):
    """Importar establiments des d'un fitxer Excel amb detecció de colors"""
    await verify_admin(authorization)
    
    print(f"📁 Import request - File: {file.filename}, Category: {category}")
    
    # Verificar que sigui un fitxer Excel
    if not file.filename:
        raise HTTPException(status_code=400, detail="No s'ha rebut cap fitxer")
    
    filename_lower = file.filename.lower()
    if not (filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls')):
        raise HTTPException(status_code=400, detail=f"El fitxer ha de ser Excel (.xlsx o .xls). Rebut: {file.filename}")
    
    try:
        from openpyxl import load_workbook
        
        # Guardar fitxer temporalment
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_path = tmp_file.name
        
        # Carregar amb openpyxl per colors
        wb = load_workbook(tmp_path)
        
        # Llegir Excel amb pandas
        df = pd.read_excel(tmp_path)
        
        imported = 0
        skipped = 0
        errors = []
        
        # Buscar columna 'Nom'
        nom_col_index = 0
        for i, col in enumerate(df.columns):
            if 'nom' in str(col).lower():
                nom_col_index = i
                break
        
        for index, row in df.iterrows():
            try:
                # Extreure dades (suporta diferents noms de columnes)
                name = str(row.get('Nom', '') or row.get('nom', '') or row.get('Nom establiment', '')).strip()
                
                # Saltar files buides
                if not name or name == 'nan' or name == '':
                    skipped += 1
                    continue
                
                # Verificar si ja existeix
                existing = await db.establishments.find_one({"name": name})
                if existing:
                    skipped += 1
                    continue
                
                # Detectar categoria pel color
                detected_category = get_category_from_cell_color(wb, index, nom_col_index)
                final_category = detected_category if detected_category else category
                
                # Preparar document
                establishment = {
                    "name": name,
                    "address": str(row.get('Adreça', '') or row.get('adreça', '')).strip() if pd.notna(row.get('Adreça')) or pd.notna(row.get('adreça')) else None,
                    "category": final_category,
                    "phone": str(row.get('Telèfon', '') or row.get('telèfon', '') or row.get('Telèfon de contacte', '')).strip() if pd.notna(row.get('Telèfon')) or pd.notna(row.get('telèfon')) or pd.notna(row.get('Telèfon de contacte')) else None,
                    "email": str(row.get('E-mail', '') or row.get('e-mail', '') or row.get('Correu electrònic', '') or row.get('correu electrònic', '')).strip() if pd.notna(row.get('E-mail')) or pd.notna(row.get('e-mail')) or pd.notna(row.get('Correu electrònic')) or pd.notna(row.get('correu electrònic')) else None,
                    "website": str(row.get('Adreça web', '')).strip() if pd.notna(row.get('Adreça web')) else None,
                    "facebook": str(row.get('Facebook', '')).strip() if pd.notna(row.get('Facebook')) else None,
                    "description": str(row.get('Descripció', '')).strip() if pd.notna(row.get('Descripció')) else None,
                    "latitude": float(row.get('Latitud')) if pd.notna(row.get('Latitud')) and row.get('Latitud') != '' else None,
                    "longitude": float(row.get('Longitud')) if pd.notna(row.get('Longitud')) and row.get('Longitud') != '' else None,
                    "created_at": datetime.utcnow(),
                }
                
                # Netejar valors 'nan'
                establishment = {k: v for k, v in establishment.items() if v and str(v) != 'nan'}
                
                # Insertar
                await db.establishments.insert_one(establishment)
                imported += 1
                
            except Exception as e:
                errors.append(f"Fila {index}: {str(e)}")
                continue
        
        # Eliminar fitxer temporal
        os.unlink(tmp_path)
        
        return {
            "success": True,
            "imported": imported,
            "skipped": skipped,
            "errors": errors[:10] if errors else []
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processant l'Excel: {str(e)}")

@admin_router.get("/stats")
async def get_stats(authorization: str = Header(None)):
    """Obtenir estadístiques generals"""
    await verify_admin(authorization)
    
    stats = {
        "users": await db.users.count_documents({}),
        "establishments": await db.establishments.count_documents({}),
        "offers": await db.offers.count_documents({}),
        "events": await db.events.count_documents({}),
        "news": await db.news.count_documents({}),
        "gift_cards": await db.gift_cards.count_documents({}),
        "active_gift_cards": await db.gift_cards.count_documents({"status": "active"}),
        "ticket_scans": await db.ticket_scans.count_documents({})
    }
    
    return stats


# ============================================
# ENDPOINTS PER GESTIÓ DE MARCADORS (TAGS)
# ============================================

@admin_router.get("/tags")
async def list_all_tags(authorization: str = Header(None)):
    """
    Llistar tots els marcadors disponibles amb estadístiques
    """
    await verify_admin(authorization)
    
    tags = await get_all_tags()
    
    return {
        "success": True,
        "tags": tags,
        "total": len(tags)
    }


@admin_router.get("/tags/{tag}/users")
async def get_tag_users(tag: str, authorization: str = Header(None)):
    """
    Obtenir tots els usuaris que han participat en activitats amb aquest marcador
    """
    await verify_admin(authorization)
    
    users = await get_users_by_tag(tag)
    stats = await get_participation_stats(tag)
    
    return {
        "success": True,
        "tag": tag,
        "users": users,
        "stats": stats
    }


@admin_router.get("/tags/{tag}/export")
async def export_tag_users(tag: str, authorization: str = Header(None)):
    """
    Exportar usuaris d'un marcador a Excel
    """
    await verify_admin(authorization)
    
    users = await get_users_by_tag(tag)
    
    # Crear Excel amb openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = f"Usuaris - {tag}"
    
    # Headers
    headers = ["Nom", "Email", "Telèfon", "Total Participacions", "Primera Participació", "Última Participació"]
    ws.append(headers)
    
    # Dades
    for user in users:
        row = [
            user.get("name", ""),
            user.get("email", ""),
            user.get("phone", ""),
            user.get("total_participations", 0),
            user.get("first_participation", "").strftime("%d/%m/%Y %H:%M") if user.get("first_participation") else "",
            user.get("last_participation", "").strftime("%d/%m/%Y %H:%M") if user.get("last_participation") else ""
        ]
        ws.append(row)
    
    # Guardar a memòria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Retornar com a fitxer descarregable
    filename = f"usuaris_{tag}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@admin_router.get("/tags/stats")
async def get_tags_stats(authorization: str = Header(None)):
    """
    Obtenir estadístiques globals de marcadors
    """
    await verify_admin(authorization)
    
    stats = await get_participation_stats()
    
    return {
        "success": True,
        "stats": stats
    }



@admin_router.get("/establishments/export-emails")
async def export_establishments_emails(authorization: str = Header(None)):
    """
    Exportar noms i correus de tots els establiments associats a Excel
    """
    await verify_admin(authorization)
    
    # Obtenir tots els establiments amb email
    establishments = await db.establishments.find(
        {"email": {"$exists": True, "$ne": None, "$ne": ""}}
    ).to_list(length=None)
    
    # Crear Excel amb openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Correus Establiments"
    
    # Headers
    headers = ["Nom", "Correu Electrònic"]
    ws.append(headers)
    
    # Dades
    for establishment in establishments:
        row = [
            establishment.get("name", ""),
            establishment.get("email", "")
        ]
        ws.append(row)
    
    # Guardar a memòria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Retornar com a fitxer descarregable
    filename = f"establiments_correus_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@admin_router.post("/fix-hosteleria-spelling")
async def fix_hosteleria_spelling(authorization: str = Header(None)):
    """
    Endpoint temporal per corregir l'ortografia de "Hostalería" → "Hostelería"
    NOTA: Aquest endpoint es pot executar una sola vegada per corregir les dades
    """
    await verify_admin(authorization)
    
    corrected_count = 0
    establishments_corrected = []
    
    # Buscar tots els establiments amb "Hostalería" (amb 'a')
    docs = await db.establishments.find(
        {"category": {"$regex": "Hostalería", "$options": "i"}}
    ).to_list(1000)
    
    for doc in docs:
        old_category = doc.get('category', '')
        new_category = old_category.replace('Hostalería', 'Hostelería')
        
        if old_category != new_category:
            await db.establishments.update_one(
                {"_id": doc['_id']},
                {"$set": {"category": new_category, "updated_at": datetime.utcnow()}}
            )
            corrected_count += 1
            establishments_corrected.append({
                "name": doc.get('name'),
                "old": old_category,
                "new": new_category
            })
    
    return {
        "success": True,
        "corrected_count": corrected_count,
        "message": f"S'han corregit {corrected_count} establiments",
        "establishments": establishments_corrected[:10]  # Mostrar només els primers 10
    }


@admin_router.delete("/users/delete-by-email/{email}")
async def delete_user_by_email(email: str, authorization: str = Header(None)):
    """
    Endpoint temporal per esborrar un usuari per email
    Útil quan la interfície no funciona
    """
    await verify_admin(authorization)
    
    # Buscar l'usuari
    user = await db.users.find_one({"email": email})
    
    if not user:
        raise HTTPException(status_code=404, detail=f"Usuari amb email {email} no trobat")
    
    user_id = user['_id']
    user_name = user.get('name', 'Sense nom')
    
    # Si l'usuari té establiments assignats, desassignar-los
    establishments = await db.establishments.update_many(
        {"owner_id": user_id},
        {"$set": {"owner_id": None, "updated_at": datetime.utcnow()}}
    )
    
    # Esborrar l'usuari
    result = await db.users.delete_one({"_id": user_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=500, detail="No s'ha pogut esborrar l'usuari")
    
    return {
        "success": True,
        "message": f"Usuari {user_name} ({email}) esborrat correctament",
        "establishments_updated": establishments.modified_count
    }



# Neteja de descripcions HTML
@admin_router.get("/cleanup-descriptions")
async def cleanup_descriptions():
    """Elimina etiquetes HTML de les descripcions dels establiments"""
    import re
    
    # Obtenir tots els establiments
    establishments = await db.establishments.find({}).to_list(length=None)
    
    updated_count = 0
    for est in establishments:
        description = est.get("description", "")
        if description and ("<p>" in description or "&lt;p&gt;" in description or "<br>" in description):
            # Netejar HTML
            clean_desc = re.sub(r'<[^>]+>', '', description)
            clean_desc = clean_desc.replace('&lt;p&gt;', '').replace('&lt;/p&gt;', '')
            clean_desc = clean_desc.replace('&nbsp;', ' ')
            clean_desc = re.sub(r'\s+', ' ', clean_desc).strip()
            
            if clean_desc != description:
                await db.establishments.update_one(
                    {"_id": est["_id"]},
                    {"$set": {"description": clean_desc}}
                )
                updated_count += 1
    
    return {
        "success": True,
        "message": f"S'han netejat {updated_count} descripcions d'establiments",
        "total_checked": len(establishments),
        "updated": updated_count
    }



# Endpoint d'estadístiques generals
@admin_router.get("/statistics")
async def get_statistics(authorization: str = Header(None)):
    """Obtenir estadístiques generals de l'aplicació"""
    await verify_admin(authorization)
    
    from datetime import datetime, timedelta
    from dateutil.relativedelta import relativedelta
    
    now = datetime.utcnow()
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    start_of_quarter = now.replace(month=((now.month - 1) // 3) * 3 + 1, day=1, hour=0, minute=0, second=0, microsecond=0)
    start_of_year = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    last_month = start_of_month - timedelta(days=1)
    start_of_last_month = last_month.replace(day=1)
    
    # Estadístiques d'usuaris
    total_users = await db.users.count_documents({})
    users_this_month = await db.users.count_documents({"created_at": {"$gte": start_of_month}})
    users_last_month = await db.users.count_documents({
        "created_at": {"$gte": start_of_last_month, "$lt": start_of_month}
    })
    users_this_quarter = await db.users.count_documents({"created_at": {"$gte": start_of_quarter}})
    users_this_year = await db.users.count_documents({"created_at": {"$gte": start_of_year}})
    
    # Calcular creixement mensual
    monthly_growth = 0
    if users_last_month > 0:
        monthly_growth = round(((users_this_month - users_last_month) / users_last_month) * 100, 1)
    
    # Estadístiques d'establiments
    total_establishments = await db.establishments.count_documents({})
    active_establishments = await db.establishments.count_documents({"is_active": True})
    
    # Estadístiques d'esdeveniments
    total_events = await db.events.count_documents({})
    active_events = await db.events.count_documents({
        "valid_until": {"$gte": now}
    })
    upcoming_events = await db.events.count_documents({
        "valid_from": {"$gte": now}
    })
    
    # Estadístiques de promocions
    total_promotions = await db.promotions.count_documents({})
    approved_promotions = await db.promotions.count_documents({"status": "approved"})
    pending_promotions = await db.promotions.count_documents({"status": "pending"})
    
    # Estadístiques de participació
    total_participations = await db.participations.count_documents({})
    participations_this_month = await db.participations.count_documents({
        "created_at": {"$gte": start_of_month}
    })
    
    # Usuaris actius (amb almenys una participació)
    active_users_pipeline = [
        {"$group": {"_id": "$user_id"}},
        {"$count": "total"}
    ]
    active_users_result = await db.participations.aggregate(active_users_pipeline).to_list(length=1)
    active_users = active_users_result[0]["total"] if active_users_result else 0
    
    # Percentatge de participació
    participation_rate = round((active_users / total_users * 100), 1) if total_users > 0 else 0
    
    # Top 5 esdeveniments per participació
    top_events_pipeline = [
        {"$match": {"activity_type": "event"}},
        {"$group": {"_id": "$activity_id", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 5}
    ]
    top_events_data = await db.participations.aggregate(top_events_pipeline).to_list(length=5)
    
    top_events = []
    for item in top_events_data:
        event = await db.events.find_one({"_id": ObjectId(item["_id"])}) if item["_id"] else None
        if event:
            top_events.append({
                "name": event.get("title", "Desconegut"),
                "participations": item["count"]
            })
    
    # Estadístiques de sortejos
    total_raffles = await db.raffles.count_documents({})
    active_raffles = await db.raffles.count_documents({
        "end_date": {"$gte": now},
        "status": "active"
    })
    
    # Estadístiques de notícies
    total_news = await db.news.count_documents({})
    news_this_month = await db.news.count_documents({"created_at": {"$gte": start_of_month}})
    
    # Marcadors més populars
    top_tags_pipeline = [
        {"$group": {"_id": "$tag", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 5}
    ]
    top_tags_data = await db.participations.aggregate(top_tags_pipeline).to_list(length=5)
    top_tags = [{"tag": item["_id"], "count": item["count"]} for item in top_tags_data if item["_id"]]
    
    # Altes mensuals dels últims 6 mesos
    monthly_signups = []
    for i in range(5, -1, -1):
        month_start = (now - relativedelta(months=i)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        month_end = (month_start + relativedelta(months=1))
        count = await db.users.count_documents({
            "created_at": {"$gte": month_start, "$lt": month_end}
        })
        monthly_signups.append({
            "month": month_start.strftime("%b %Y"),
            "count": count
        })
    
    # Participacions per tipus d'activitat
    activity_types_pipeline = [
        {"$group": {"_id": "$activity_type", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    activity_types_data = await db.participations.aggregate(activity_types_pipeline).to_list(length=10)
    participation_by_type = {item["_id"]: item["count"] for item in activity_types_data if item["_id"]}
    
    return {
        "users": {
            "total": total_users,
            "this_month": users_this_month,
            "last_month": users_last_month,
            "this_quarter": users_this_quarter,
            "this_year": users_this_year,
            "monthly_growth": monthly_growth,
            "active_users": active_users,
            "participation_rate": participation_rate
        },
        "establishments": {
            "total": total_establishments,
            "active": active_establishments
        },
        "events": {
            "total": total_events,
            "active": active_events,
            "upcoming": upcoming_events,
            "top_events": top_events
        },
        "promotions": {
            "total": total_promotions,
            "approved": approved_promotions,
            "pending": pending_promotions
        },
        "raffles": {
            "total": total_raffles,
            "active": active_raffles
        },
        "news": {
            "total": total_news,
            "this_month": news_this_month
        },
        "participations": {
            "total": total_participations,
            "this_month": participations_this_month,
            "by_type": participation_by_type
        },
        "trends": {
            "monthly_signups": monthly_signups,
            "top_tags": top_tags
        }
    }
